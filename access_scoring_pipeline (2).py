
# dependecies
# !pip install pymupdf pdfplumber
# !pip install rapidfuzz
# !pip install rank-bm25
# !pip install sentence-transformers faiss-cpu -q
# !pip install rank-bm25
# !pip install groq


from openpyxl import load_workbook
import re
import pandas as pd
import numpy as np

import fitz
import pdfplumber
import re
import json
import numpy as np
import faiss
import time
import torch

from collections import Counter
from transformers import AutoModelForCausalLM

from sentence_transformers import (
    SentenceTransformer,
    CrossEncoder
)

from rank_bm25 import BM25Okapi

from groq import Groq


from openpyxl import load_workbook

#============Inputs=================

EXCEL_PATH = ""
PDF_FOLDER = ""
SHEET_NAME = "Submissions"
GROQ_API_KEY = ""
#Current model used is Llama 3.3 70B
#if test file has more than 7-8 files to be processed at once used llama 8b model


# =========================================================
# AGE CRITERIA — RETRIEVAL PROMPT TEMPLATE
# =========================================================

AGE_RETRIEVAL_PROMPT = """
{DRUG_NAME} psoriasis age criteria

Target retrieval toward chunks where:
- {DRUG_NAME}
AND
- Psoriasis OR PsO OR plaque psoriasis

appear in the same chunk or nearby text.

Strongly prioritize chunks containing:
- {DRUG_NAME}
- Psoriasis
- PsO
- plaque psoriasis
- age ≥
- age requirement
- years and older
- adult patients
- FDA approved indication
- FDA labelled age
- biologic criteria
- prior authorization criteria
- approval criteria
- universal psoriasis criteria
- All other drugs

Also prioritize:
- plaque psoriasis indications
- psoriasis approval criteria
- psoriasis biologic criteria


Prioritize:
1. Drug-specific psoriasis criteria
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal psoriasis criteria
"""

STEP_THERAPY_RETRIEVAL_PROMPT = """
Retrieve chunks related to STEP THERAPY REQUIREMENTS for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque psoriasis

Priority retrieval targets:

- {DRUG_NAME}
- psoriasis
- plaque psoriasis
- step therapy
- prior authorization
- psoriasis biologic criteria
- psoriasis coverage criteria
- psoriasis approval criteria
- systemic therapy
- prerequisite therapy
- previous therapy
- prior therapy
- conventional systemic therapy
- non-biologic systemic therapy
- preferred product
- preferred biologic
- TNF inhibitor
- formulary exception
- bypass
- waiver
- prerequisite drug
- required drug history
- All other drugs
- step therapy requirements
- required previous therapies
- therapies that must be tried before approval
- therapies that must be failed before approval
- required treatment history



Therapy requirement keywords:

- phototherapy
- {DRUG_NAME}
- biologic
- targeted synthetic drug
- fail
- failed
- inadequate response
- intolerance
- unable to take
- must try
- must fail
- after trial of
- previous biologic
"""

BRANDED_STEP_COUNT_RETRIEVAL_PROMPT = """
Retrieve chunks related to BRANDED / BIOLOGIC STEP THERAPY requirements for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or plaque Psoriasis

Focus on chunks containing:
- {DRUG_NAME}
- biologic step therapy
- branded step therapy
- preferred product requirements
- non-preferred product requirements
- TNF inhibitor
- CAM antagonist
- targeted synthetic drug
- prior biologic use
- failure of biologic therapy
- trial and failure
- inadequate response
- required previous biologic
- step through preferred product
- preferred formulary product
- requested product
- approval criteria
- initial authorization
- plaque psoriasis criteria
- moderate to severe plaque psoriasis
- preferred biosimilar
- reference product
- non-preferred biologic
- formulary
- step through
- bypass criteria

Prioritize:
1. Drug-specific PsO criteria
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal PsO biologic criteria
4. Preferred vs non-preferred product requirements
5. Drug-class level requirements applicable to {DRUG_NAME}

"""

GENERIC_STEP_COUNT_RETRIEVAL_PROMPT = """
Retrieve chunks related to STEP THERAPY requirements for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on chunks containing:
- step therapy
- trial and failure
- inadequate response
- contraindication
- prerequisite therapy
- prior treatment
- required treatment history
- topical therapy
- biologic step requirements
- targeted synthetic drug requirements
- plaque psoriasis criteria
- universal biologic criteria
- approval criteria
- initial authorization criteria
- generic treatments
- non biologic
- topical corticosteroids
- vitamin D analogs
- topical agents

Prioritize:
1. Drug-specific PsO criteria
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal PsO biologic criteria
4. Preferred vs non-preferred product step requirements
"""

PHOTOTHERAPY_STEP_RETRIEVAL_PROMPT = """
Retrieve chunks related to PHOTOTHERAPY step requirements for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on chunks containing:
- phototherapy
- UVB
- PUVA
- ultraviolet therapy
- light therapy
- trial and failure
- inadequate response
- intolerance
- contraindication
- prerequisite therapy
- required previous treatment
- systemic therapy
- psoriasis biologic criteria
- psoriasis approval criteria
- psoriasis initial authorization
- initial authorization
- plaque psoriasis criteria
- moderate to severe plaque psoriasis

Prioritize:
1. Drug-specific PsO criteria phototherapy
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal PsO biologic criteria
4. Step therapy sections, phototherapy

"""

TB_TEST_RETRIEVAL_PROMPT = """
Retrieve chunks related to TUBERCULOSIS (TB) TESTING requirements for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on chunks containing:
- TB test
- tuberculosis test
- latent tuberculosis
- TB screening
- tuberculosis screening
- negative TB test
- tuberculin skin test
- latent TB infection
- active tuberculosis
- prior to therapy
- before initiation
- before starting treatment
- biologic criteria
- approval criteria
- Universal criteria
- General criteria
- prior authorization criteria
- initial authorization
- monitoring requirements
- All other drugs

Prioritize:
1. Drug-specific PsO criteria
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal PsO biologic criteria
4. Safety monitoring requirements

"""

INITIAL_AUTH_DURATION_RETRIEVAL_PROMPT = """
Retrieve chunks related to INITIAL AUTHORIZATION DURATION for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on chunks containing:
- authorization duration
- Approval Length
- initial authorization
- approval duration
- coverage period
- coverage criteria
- authorization period
- length of approval
- duration of approval
- authorization may be granted
- approval may be granted
- initial coverage
- PA duration
- prior authorization duration
- authorization timeframe
- plaque psoriasis criteria
- moderate to severe plaque psoriasis
- All other drugs

Prioritize:
1. Drug-specific PsO authorization duration
2. Criteria explicitly applicable to {DRUG_NAME}
3. Universal PsO biologic authorization duration

"""

REAUTHORIZATION_RETRIEVAL_PROMPT = """
Retrieve chunks related to REAUTHORIZATION / CONTINUATION CRITERIA for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on:

{DRUG_NAME}

- reauthorization
- reauthorization criteria
- reauthorization duration
- continuation criteria
- continuation approval
- continuation of therapy
- renewal criteria
- renewal requirements
- renewal duration
- continued coverage
- continued therapy
- Continuation requests
- Continuation requests
- maintenance therapy
- maintenance authorization
- approval length
- authorization period
- coverage period
- duration of approval
- response to therapy
- clinical response
- continued clinical benefit
- improvement from baseline
- reduction in BSA

Look for durations such as:
- 3 months
- 6 months
- 12 months
- 24 months
- annually
- yearly
- every 6 months

Prioritize:

1. Drug-specific PsO continuation criteria
2. Universal PsO continuation criteria
3. General biologic continuation criteria

Return chunks containing:
- continuation requirements
- renewal requirements
- approval duration
- renewal duration
- reauthorization duration
- clinical response requirements
"""

SPECIALIST_TYPES_RETRIEVAL_PROMPT = """
Retrieve chunks related to specialist requirements for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on:
{DRUG_NAME}
- specialist requirements
- prescriber requirements
- prescribing physician
- specialist physician
- provider specialty
- treatment managed by
- consultation with specialist
- dermatologist
- rheumatologist
- gastroenterologist
- immunologist
- specialist supervision
- managed by specialist
- Prescriber Specialties

Look for statements such as:
- prescribed by
- Psoriasis (PsO) or Plaque Psoriasis
- under supervision of
- consultation with
- specialist required
- treatment initiated by
- board-certified specialist

Prioritize:
1. {DRUG_NAME} PsO specialist requirements
2. Criteria explicitly applicable to the target drug
3. Universal PsO specialist criteria
4. General biologic specialist criteria
"""

QUANTITY_LIMITS_RETRIEVAL_PROMPT = """
Retrieve chunks related to quantity limits for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

Focus on:
{DRUG_NAME}
- quantity limit
- dispensing limit
- maximum quantity
- limit per days supply
- vials per days
- syringes per days
- pens per days
- exception limit
- quantity level limit

Look specifically for:
- quantity limit
- quantity level limit

Prioritize:
1. {DRUG_NAME} quantity limit sections
2. Criteria explicitly applicable to the target drug
3. Universal quantity limit sections for Psoriasis PsO biologics

"""

QUANTITY_LIMITS_EXTRACTION_PROMPT = """
You are an expert medical policy analyst.

Your task is to extract ONLY the FINAL APPLICABLE QUANTITY LIMITS for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO)

====================================================================
IMPORTANT CONTEXT
====================================================================

Retrieved chunks may contain:
- quantity limit sections
- dosing sections
- dosage limits
- administration instructions
- biologic criteria
- multiple drugs
- multiple diseases

You must determine ONLY the FINAL applicable quantity limits for:
- {DRUG_NAME}
AND
- Psoriasis (PsO)

-nearest subsection heading is the latest heading if multiple headings exist

====================================================================
BUSINESS RULE
====================================================================

Quantity Limits means ONLY:
explicitly stated quantity limit restrictions.

Examples:
- quantity limit
- quantity level limit
- exception limit
- maximum quantity allowed

IMPORTANT:

Do NOT capture:
- dosage instructions
- dosing schedules
- dosing limits
- administration frequency

ONLY extract content explicitly presented as:
- quantity limits or Quantity limits:
OR
- quantity level limits

===============================================================
DOCUMENT-SPECIFIC POLICY RULE

Some policies are dedicated to a single therapy.

If:
- the document title, policy title, or major heading clearly identifies {DRUG_NAME}
- no other drug-specific policy sections are present,but indications are mentioned at many places
AND
- the retrieved chunks appear to belong to the same policy document

then you may treat the document as a dedicated {DRUG_NAME} policy.

If you have a mixture of drug name and indications in the do not use it

In this situation:

- criteria do NOT need to repeatedly mention {DRUG_NAME}
- determine applicability primarily using indication matching
- continue to exclude criteria belonging to unrelated indications

Examples:
- Psoriatic Arthritis
- Generalized Pustular Psoriasis (GPP)
- Crohn's Disease
- Ulcerative Colitis

Drug applicability may be inferred from the document context,
but indication applicability must still be explicitly established.

Use this criteria very cautiously

====================================================================
PRIORITY ORDER
====================================================================

PRIORITY 1
------------------------------------------------

Drug-specific quantity limits for:
- {DRUG_NAME}


PRIORITY 2
------------------------------------------------

Criteria explicitly applicable to:
- {DRUG_NAME}

Examples:
- applies to the following drugs
- applicable therapies
- for all biologics below


PRIORITY 3
------------------------------------------------

Universal quantity limit sections applicable to PsO biologics.
Use ONLY if:
- no drug-specific age exists
AND
- the section clearly applies to psoriasis therapies

PRIORITY 4
------------------------------------------------

If there exist a general heading in document related to Quantity limit criteria then it can be used if its written in general format
Just try to ensure that complete paragraph is retrieved by looking at chunk structure
if you feel its complete and generally written (there can be exceptions given though look at them too) then you can give that as criteria If youn dont think its general format just return "NA"

Priority 5
--------------------------------------------------------
When {DRUG_NAME} is not explicitly repeated within a criterion section, use document-level context. If the policy is clearly a drug-specific policy for {DRUG_NAME} and the section is discussing the target indication (e.g., Plaque Psoriasis/PsO), assume the criteria apply to {DRUG_NAME} unless the text explicitly states otherwise. Do not reject a criterion solely because {DRUG_NAME} is absent from that specific chunk.
====================================================================
EXTRACTION RULES
====================================================================

1. Extract ONLY explicit quantity limit statements.

2. Preserve:
- units
- vial/ syringe counts
- day supply windows
- exception limits
- bullet structure

3. Include:
- quantity level limit
- exception limit
- quantity allowed per time period

4. Ignore:
- dosing frequency
- induction dosing
- maintenance dosing
- administration instructions
- dose escalation language

5. Preserve wording as closely as possible.

6. If multiple quantity limits exist for the target drug,
include all applicable quantity limits.

7. Use ONLY provided text.

====================================================================
INDICATION MATCHING RULE
====================================================================

The same drug may appear under MULTIPLE indications.

Examples:
- Plaque Psoriasis
- Generalized Pustular Psoriasis (GPP)
- Psoriatic Arthritis (PsA)
- Crohn's Disease
- Ulcerative Colitis

You must extract criteria for:
- Psoriasis (PsO)
OR
- plaque psoriasis

generalized pustular psoriasis these kind of indication are different from plaque psoriasis or PsO

====================================================================
NA RULE
====================================================================

Return "NA" if:
- no explicit quantity limit exists
OR
- only dosing information exists
OR
- only dosage limits exist
-------------------------------------------------------------------
IMPORTANT OUTPUT RULES

Return ONLY valid JSON.

Do NOT:
- add explanations
- add markdown
- add text before JSON
- add text after JSON


====================================================================
OUTPUT FORMAT
====================================================================

Return ONLY valid JSON.

{{
    "Quantity Limits": "<final extracted quantity limits or NA>"
}}



====================================================================
POLICY TEXT
====================================================================

{quantity}
"""

Eligibility_Extraction_prompt="""You are an expert medical policy analyst.

Your task is to extract the FINAL APPLICABLE values for the following attributes for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

====================================================================
IMPORTANT CONTEXT
=================

The policy text provided below may contain THREE SEPARATE RETRIEVED SECTIONS:

1. AGE RETRIEVED SECTION

   * Retrieved specifically for age eligibility criteria

2. SPECIALIST RETRIEVED SECTION

   * Retrieved specifically for specialist/prescriber requirements

3. TB RETRIEVED SECTION

   * Retrieved specifically for tuberculosis testing requirements

These sections may:

* overlap
* contain duplicate content
* contain partial chunks
* contain general criteria
* contain drug-specific criteria
* contain indication-specific criteria
* contain biologic criteria
* contain approval criteria
* contain continuation criteria
* contain multiple diseases
* contain multiple drugs

You MUST use information from ALL retrieved sections when determining the final output.

A criterion found in any retrieved section may be used if it satisfies the applicability rules below.

Nearest subsection heading is considered the active heading when multiple headings exist.

====================================================================
TARGET EXTRACTION
=================

Extract ONLY the FINAL APPLICABLE values for:

1. Age
2. Specialist Types
3. TB Test Required

for:

* {DRUG_NAME}
  AND
* Psoriasis (PsO) / Plaque Psoriasis

====================================================================
BUSINESS RULE
=============

If policy separates:

* mild psoriasis
* moderate psoriasis
* moderate-to-severe psoriasis
* severe psoriasis
* other psoriasis subtypes

Always prioritize:

MODERATE-TO-SEVERE PLAQUE PSORIASIS / PsO CRITERIA

If both moderate-to-severe and broader/general psoriasis criteria exist:

* use moderate-to-severe psoriasis criteria

If moderate-to-severe criteria do not contain the requested attribute:

* then use applicable general psoriasis criteria.

====================================================================
APPLICABILITY RULE
==================

General sections such as:

* Prior Authorization Criteria
* Initial Authorization
* Coverage Criteria
* Approval Criteria
* Clinical Criteria
* Program Criteria
* Universal Criteria
* Biologic Criteria

should ONLY be used when the document explicitly indicates they apply to:

* {DRUG_NAME}
  OR
* the requested biologic therapy class
  OR
* plaque psoriasis therapies
  OR
* psoriasis biologics

Examples of applicability language:

* applies to the following drugs
* applicable therapies
* this policy applies to
* for all biologics below
* for plaque psoriasis
* for requested biologic
* criteria for the following agents
* applies to the requested therapy

Do NOT automatically assume a general section applies.

====================================================================
DOCUMENT-SPECIFIC POLICY RULE
=============================

Some policies are dedicated to a single therapy.

You may treat the policy as a dedicated {DRUG_NAME} policy ONLY IF:

* document title clearly identifies {DRUG_NAME}
  OR
* policy title clearly identifies {DRUG_NAME}
  OR
* major heading clearly identifies {DRUG_NAME}

AND

* no other drug-specific policy sections are present

AND

* retrieved chunks appear to belong to the same policy document

In this situation:

* criteria do not need to repeatedly mention {DRUG_NAME}
* drug applicability may be inferred from document context

HOWEVER:

Indication applicability must still be explicitly established.

Do NOT use this assumption if:

* multiple drugs appear in the retrieved text
* drug-specific sections for other therapies are present
* chunks appear to come from mixed policies

Use this rule very cautiously.

====================================================================
INDICATION MATCHING RULE
========================

The same drug may appear under multiple indications.

Examples:

* Plaque Psoriasis
* Psoriasis (PsO)
* Generalized Pustular Psoriasis (GPP)
* Psoriatic Arthritis (PsA)
* Crohn's Disease
* Ulcerative Colitis
* Hidradenitis Suppurativa

Extract criteria ONLY for:

* Plaque Psoriasis
  OR
* Psoriasis (PsO)

Do NOT use criteria that apply exclusively to:

* PsA
* GPP
* Crohn's Disease
* Ulcerative Colitis
* any other non-PsO indication

====================================================================
PRIORITY ORDER
==============

## Priority 1

Drug-specific PsO criteria for:

* {DRUG_NAME}

Examples:

* "{DRUG_NAME} for plaque psoriasis"
* "{DRUG_NAME} criteria"
* "{DRUG_NAME} may be used in adults"

Use these first.

## Priority 2

Criteria explicitly applicable to:

* {DRUG_NAME}

Examples:

* applies to the following drugs
* applicable therapies
* requested therapy
* for all biologics below

Use only when {DRUG_NAME} is clearly included.

## Priority 3

Universal PsO / Psoriasis biologic criteria.

Use only if:

* no drug-specific criterion exists for the attribute
  AND
* section clearly applies to psoriasis therapies

## Priority 4

General criteria sections.

Use only if:

* heading clearly indicates the criterion type
  AND
* the text appears complete
  AND
* the section is written as a generally applicable policy requirement

If applicability is unclear, do not use it.

## Priority 5

When {DRUG_NAME} is not explicitly repeated within a criterion section:

If the document is clearly a dedicated {DRUG_NAME} policy and the section is discussing PsO/Plaque Psoriasis, assume the criterion applies to {DRUG_NAME} unless the text explicitly states otherwise.

Do not reject a criterion solely because {DRUG_NAME} is absent from that chunk.

====================================================================
AGE EXTRACTION RULES
====================

Extract ONLY the FINAL APPLICABLE AGE requirement.

Normalize as:

* adults only → >=18 years
* adult patients → >=18 years
* age 18 years or older → >=18 years
* 12 years and older → >=12 years
* pediatric patients 6 years and older → >=6 years

## FDA LABEL RULE

Use:

"FDA labelled age"

ONLY if policy explicitly states:

* FDA approved age
* FDA labelled age
* according to FDA label
* FDA indication age
* age per FDA approval

AND

no explicit numeric age is provided.

Do NOT use external FDA knowledge.

## MULTIPLE AGE RULE

If multiple valid applicable ages exist:

Return the YOUNGEST applicable age.

Ignore ages belonging only to:

* other drugs
* other diseases
* other indications

If no applicable age exists:

Return:

"Not Specified"

====================================================================
SPECIALIST EXTRACTION RULES
===========================

Extract ONLY specialist/prescriber requirements.

Examples:

* dermatologist
* rheumatologist
* gastroenterologist
* immunologist
* specialist physician

Normalization:

* dermatology specialist → dermatologist
* rheumatology specialist → rheumatologist

If multiple specialists are acceptable:

Return all valid specialists while preserving OR/AND meaning.

Examples:

* dermatologist OR rheumatologist
* dermatologist AND rheumatologist

If policy states:

* specialist required

but specialty is not specified:

Return:

"Specialist Required - Type Not Specified"

If no applicable specialist requirement exists:

Return:

"NA"

====================================================================
TB TEST EXTRACTION RULES
========================

Determine whether TB testing is required.

Return:

"Y"

ONLY if policy explicitly requires:

* TB testing
* TB screening
* tuberculosis screening
* latent TB evaluation
* negative TB test
* tuberculosis clearance

before approval or treatment initiation.

Examples:

* negative TB test required
* tuberculosis screening required prior to therapy
* must be evaluated for latent tuberculosis

Return:

"N"

If:

* no applicable TB requirement exists
  OR
* TB testing is not required

====================================================================
IMPORTANT RULES
===============

1. Use ONLY provided text.

2. Do NOT use external medical knowledge.

3. Do NOT hallucinate missing criteria.

4. Ignore unrelated drugs.

5. Ignore unrelated indications.

6. Ignore unrelated diseases.

7. Preserve policy meaning exactly.

8. If applicability is unclear, do not use the criterion.

9. Determine Age, Specialist Types, and TB Test independently using all retrieved sections.


====================================================================
OUTPUT FORMAT
=============

Return ONLY valid JSON.

Do NOT:

* add explanations
* add markdown
* add text before JSON
* add text after JSON

{{
"Age": "<final applicable age or Not Specified>",
"Specialist Types": "<final applicable specialist types or NA>",
"TB Test required": "<Y | N>"
}}

====================================================================
POLICY TEXT
===========

Age retrieved chunks- {age}

Specialist chunks- {specialist}

TB specefic retrieved chunks- {tb}

You may use information from ANY retrieval section.

A criterion may appear in a section retrieved for another attribute.

Do not restrict extraction to the section name.
Use all retrieved evidence when determining the final answer.
"""

Step_Extraction_Prompt= """You are an expert medical policy analyst.

Your task is to extract the FINAL APPLICABLE STEP THERAPY REQUIREMENTS for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

====================================================================
IMPORTANT CONTEXT
=================

The policy text provided below may contain FOUR SEPARATE RETRIEVED SECTIONS:

1. STEP THERAPY RETRIEVED SECTION

   * Retrieved specifically for step therapy requirements

2. GENERIC STEP COUNT RETRIEVED SECTION

   * Retrieved specifically for generic/non-biologic step counting

3. BRANDED STEP COUNT RETRIEVED SECTION

   * Retrieved specifically for biologic/branded step counting

4. PHOTOTHERAPY RETRIEVED SECTION

   * Retrieved specifically for phototherapy step requirements

The retrieved sections may:

* overlap
* contain duplicate text
* contain partial chunks
* contain universal criteria
* contain biologic criteria
* contain approval criteria
* contain prior authorization criteria
* contain multiple drugs
* contain multiple diseases
* contain multiple indications

You must use information from ALL retrieved sections when determining the final outputs.

Nearest subsection heading is considered the active heading when multiple headings exist.

====================================================================
TARGET EXTRACTION
=================

Determine:

1. Step Therapy Requirements Documented in Policy
2. Number of Steps through Generic
3. Number of Steps through Brands
4. Step through-Phototherapy

for:

* {DRUG_NAME}
  AND
* Psoriasis (PsO) / Plaque Psoriasis

====================================================================
BUSINESS RULE
=============

Extract ALL applicable step therapy language including:

* prerequisite therapies
* failure requirements
* intolerance requirements
* contraindication requirements
* prior biologic requirements
* targeted synthetic requirements
* trial requirements
* step-through requirements

If policy distinguishes:

* mild psoriasis
* moderate psoriasis
* moderate-to-severe psoriasis
* severe psoriasis
* other psoriasis subtypes

Always prioritize:

MODERATE-TO-SEVERE PLAQUE PSORIASIS / PsO CRITERIA

If moderate-to-severe criteria do not contain the requested attribute:
use applicable general psoriasis criteria.

====================================================================
APPLICABILITY RULE
==================

General sections such as:

* Prior Authorization Criteria
* Initial Authorization
* Approval Criteria
* Coverage Criteria
* Clinical Criteria
* Program Criteria
* Universal Criteria
* Biologic Criteria
* Step Therapy
* Systemic Therapy Criteria

should ONLY be used if they explicitly apply to:

* {DRUG_NAME}
  OR
* psoriasis therapies
  OR
* requested biologic therapies
  OR
* requested therapy class

Examples:

* applies to the following drugs
* applicable therapies
* this policy applies to
* criteria for the following agents
* for plaque psoriasis
* for all biologics
* for requested therapy

Do NOT automatically assume applicability.

====================================================================
DOCUMENT-SPECIFIC POLICY RULE
=============================

Some policies are dedicated to a single therapy.

You may treat the policy as dedicated to {DRUG_NAME} ONLY IF:

* document title identifies {DRUG_NAME}
  OR
* policy title identifies {DRUG_NAME}
  OR
* major heading identifies {DRUG_NAME}

AND

* no other drug-specific sections are present

AND

* retrieved chunks appear to belong to the same policy document

In this situation:

* drug applicability may be inferred
* indication applicability must still be established

Do NOT use this assumption when:

* multiple drugs appear
* mixed policy content appears
* other drug-specific sections exist

Use cautiously.

====================================================================
INDICATION MATCHING RULE
========================

Extract criteria ONLY for:

* Plaque Psoriasis
  OR
* Psoriasis (PsO)

Do NOT use criteria applicable exclusively to:

* Psoriatic Arthritis
* Generalized Pustular Psoriasis (GPP)
* Crohn's Disease
* Ulcerative Colitis
* Hidradenitis Suppurativa
* other non-PsO indications

====================================================================
PRIORITY ORDER
==============

## Priority 1

Drug-specific PsO criteria for:

* {DRUG_NAME}

Examples:

* Before approval of {DRUG_NAME}
* Member must have failed
* {DRUG_NAME} criteria

## Priority 2

Criteria explicitly applicable to:

* {DRUG_NAME}
* requested biologic class

## Priority 3

Universal PsO criteria.

Use when:

* no drug-specific criteria exist

## Priority 4

General biologic or step therapy criteria.

Use only when:

* no drug-specific criteria exist
  AND
* no psoriasis-specific criteria exist

## Priority 5

General headings discussing step therapy.

Use only if:

* section appears complete
* applicability is reasonably clear

## Priority 6

If policy is clearly dedicated to {DRUG_NAME}, do not reject a criterion solely because {DRUG_NAME} is absent from the specific chunk.

====================================================================
STEP THERAPY EXTRACTION RULES
=============================

Extract all applicable:

* required prior therapies
* treatment failures
* inadequate response requirements
* intolerance requirements
* contraindication requirements
* biologic prerequisites
* targeted synthetic prerequisites
* must try
* must fail
* after inadequate response
* after trial of
* unable to take

Preserve:

* AND logic
* OR logic
* nested logic
* combination logic

Do NOT simplify logic.

If policy explicitly states:

* no step therapy required
* exempt from step therapy
* no prerequisite therapy required

return that statement.

If no applicable step therapy exists:

return:

"Not Specified"

====================================================================
GENERIC STEP COUNT RULES
========================

Generic steps include:

* methotrexate
* cyclosporine
* acitretin
* topical corticosteroids
* vitamin D analogs
* topical therapies
* conventional systemic therapies

Topical therapies count as GENERIC.

Do NOT count:

* biologics
* targeted synthetic drugs
* branded therapies
* phototherapy

COUNTING LOGIC

1. Combine all applicable criteria.

2. Treat combined criteria as AND requirements.

3. For OR pathways:
   choose the LEAST RESTRICTIVE pathway.

Examples:

methotrexate OR cyclosporine
→ 1

methotrexate AND cyclosporine
→ 2

methotrexate OR adalimumab
→ 1

If no generic steps are required:

return:

"NA"

====================================================================
BRANDED STEP COUNT RULES
========================

Count:

* biologics
* branded products
* preferred biologics
* biosimilars
* targeted synthetic drugs
* class-level biologic requirements

Examples:

* adalimumab products
* ustekinumab products
* TNF inhibitors
* IL inhibitors
* preferred biologics

A class-level biologic requirement counts as:

1 branded step

COUNTING LOGIC

1. Combine all applicable criteria.

2. Treat combined criteria as AND requirements.

3. For OR pathways:
   choose least restrictive pathway.

Examples:

adalimumab OR ustekinumab
→ 1

adalimumab AND ustekinumab
→ 2

TNF inhibitor OR ustekinumab
→ 1

Do NOT count:

* topical therapies
* methotrexate
* cyclosporine
* acitretin
* phototherapy

If no branded steps exist:

return:

"NA"

====================================================================
PHOTOTHERAPY RULES
==================

Phototherapy includes:

* UVB
* PUVA
* ultraviolet light therapy
* light therapy

Return:

"Yes"

ONLY if phototherapy is mandatory.

Examples:

* must fail phototherapy
* inadequate response to phototherapy required

Return:

"No"

If:

* phototherapy is optional
  OR
* phototherapy appears in OR logic

Examples:

* phototherapy OR methotrexate
* phototherapy OR systemic therapy

Return:

"NA"

ONLY if no applicable criteria exist.

Phototherapy NEVER contributes to:

* generic count
* branded count

====================================================================
IMPORTANT RULES
===============

1. Use ONLY provided text.

2. Do NOT use external knowledge.

3. Do NOT hallucinate.

4. Ignore unrelated drugs.

5. Ignore unrelated indications.

6. Preserve policy logic exactly.

7. If applicability is unclear, do not use the criterion.

8. Determine all four outputs independently.


====================================================================
OUTPUT FORMAT
=============

Return ONLY valid JSON.

{{
"Step Therapy Requirements Documented in Policy": "<final extracted value or Not Specified>",
"Number of Steps through Generic": "<number or NA>",
"Number of Steps through Brands": "<number or NA>",
"Step through-Phototherapy": "<Yes | No | NA>"
}}

====================================================================
POLICY TEXT
===========

Step Therapy Requirements chunks={step}
Number of Steps through Generic chunks={generic}
Number of steps through Brands={brand}
Phototherapy chunks={phototherapy}

You may use information from ANY retrieval section.

A criterion may appear in a section retrieved for another attribute.

Do not restrict extraction to the section name.
Use all retrieved evidence when determining the final answer.
"""

Authorization_Extraction_Prompt="""You are an expert medical policy analyst.

Your task is to determine the FINAL APPLICABLE AUTHORIZATION INFORMATION for:

Drug:
{DRUG_NAME}

Indication:
Psoriasis (PsO) or Plaque Psoriasis

====================================================================
IMPORTANT CONTEXT
=================

The policy text provided below may contain TWO SEPARATE RETRIEVED SECTIONS:

1. INITIAL AUTHORIZATION RETRIEVED SECTION

   * Retrieved specifically for initial authorization duration

2. REAUTHORIZATION RETRIEVED SECTION

   * Retrieved specifically for continuation, renewal, maintenance, or reauthorization criteria

Retrieved sections may contain:

* prior authorization criteria
* approval duration
* initial authorization duration
* continuation criteria
* reauthorization criteria
* renewal criteria
* maintenance therapy requirements
* continuation requests
* approval periods
* multiple drugs
* multiple diseases
* multiple indications

You must use information from ALL retrieved sections.

Nearest subsection heading is considered the active heading when multiple headings exist.

====================================================================
TARGET EXTRACTION
=================

Determine:

1. Initial Authorization Duration(in-months)

2. Reauthorization Required

3. Reauthorization Duration(in-months)

4. Reauthorization Requirements Documented in Policy

for:

* {DRUG_NAME}
  AND
* Psoriasis (PsO) / Plaque Psoriasis

====================================================================
BUSINESS DEFINITIONS
====================

## FIELD 1

Initial Authorization Duration(in-months)

Definition:

Length of coverage granted after initial approval.

Examples:

* 3 Months
* 6 Months
* 12 Months
* 24 Months

Return:

* numeric month value

OR

* Unspecified

Examples:

* 12
* 6
* Unspecified

Ignore:

* renewal duration
* continuation duration
* maintenance approval duration

---

## FIELD 2

Reauthorization Required

Definition:

Whether reassessment and renewal approval are required after initial authorization expires.

Return:

"Yes"

if either:

* reauthorization duration exists

OR

* continuation criteria exist

OR

* renewal criteria exist

OR

* reauthorization criteria exist

Return:

"No"

ONLY if:

* no continuation criteria
  AND
* no reauthorization criteria
  AND
* no renewal criteria

exist.

---

## FIELD 3

Reauthorization Duration(in-months)

Definition:

Approval duration granted after successful renewal.

Examples:

* 6 Months
* 12 Months
* 24 Months

Return:

* numeric month value

OR

* Unspecified

OR

* NA

Rules:

If Reauthorization Required = Yes

and duration is documented:

return numeric months.

Convert:

* 1 year → 12
* 2 years → 24

If Reauthorization Required = Yes

but duration is not documented:

return:

"Unspecified"

If Reauthorization Required = No

return:

"NA"

---

## FIELD 4

Reauthorization Requirements Documented in Policy

Definition:

Continuation criteria required to maintain approval.

Examples:

* positive clinical response
* continued benefit
* reduction in body surface area
* maintained disease control
* improvement from baseline
* stable disease

Return:

* exact applicable continuation requirements

OR

"Unspecified"

if continuation approval is clearly required but requirements are not documented.

OR

"NA"

if reauthorization is not required.

====================================================================
INDICATION RULE
===============

Extract criteria ONLY for:

* Psoriasis (PsO)
  OR
* Plaque Psoriasis

Do NOT use criteria belonging exclusively to:

* Psoriatic Arthritis
* Generalized Pustular Psoriasis (GPP)
* Crohn's Disease
* Ulcerative Colitis
* Rheumatoid Arthritis
* Ankylosing Spondylitis
* Hidradenitis Suppurativa

GPP is NOT equivalent to Plaque Psoriasis.

====================================================================
SEVERITY RULE
=============

If policy separates:

* mild psoriasis
* moderate psoriasis
* moderate-to-severe psoriasis
* severe psoriasis

Always prioritize:

MODERATE-TO-SEVERE PSORIASIS CRITERIA

If moderate-to-severe criteria do not exist:

use applicable plaque psoriasis criteria.

====================================================================
APPLICABILITY RULE
==================

General sections such as:

* Prior Authorization Criteria
* Initial Authorization
* Approval Duration
* Continuation Criteria
* Reauthorization Criteria
* Renewal Criteria
* Maintenance Criteria
* Coverage Criteria
* Clinical Criteria

should ONLY be used if they explicitly apply to:

* {DRUG_NAME}
  OR
* psoriasis therapies
  OR
* requested therapy class

Examples:

* applies to the following drugs
* this policy applies to
* applicable therapies
* for plaque psoriasis
* for all biologics
* requested therapy

Do NOT assume applicability automatically.

====================================================================
DOCUMENT-SPECIFIC POLICY RULE
=============================

Some policies are dedicated to a single therapy.

You may treat the policy as dedicated to {DRUG_NAME} ONLY IF:

* document title identifies {DRUG_NAME}
  OR
* policy title identifies {DRUG_NAME}
  OR
* major heading identifies {DRUG_NAME}

AND

* no other drug-specific sections are present

AND

* retrieved chunks appear to belong to the same policy document

In this situation:

* drug applicability may be inferred

However:

indication applicability must still be established.

Do NOT use this assumption when:

* multiple drugs appear
* mixed policy content appears
* other drug-specific sections exist

Use cautiously.

====================================================================
PRIORITY ORDER
==============

## Priority 1

Drug-specific PsO authorization criteria
for {DRUG_NAME}

## Priority 2

Criteria explicitly applicable to:

* {DRUG_NAME}

## Priority 3

Universal PsO authorization criteria.

Use only when:

* no drug-specific criteria exist

AND

* section clearly applies to psoriasis therapies.

## Priority 4

General authorization or continuation sections.

Use only if:

* section appears complete
* applicability is reasonably clear

Otherwise ignore.

## Priority 5

If policy is clearly dedicated to {DRUG_NAME}, do not reject a criterion solely because {DRUG_NAME} is absent from the specific chunk.

====================================================================
INITIAL AUTHORIZATION RULES
===========================

Extract ONLY INITIAL authorization duration.

Ignore:

* renewal duration
* continuation duration
* maintenance duration
* reauthorization duration

Examples:

authorization for 12 months
→ 12

approval granted for 6 months
→ 6

If PsO approval exists but duration is not stated:

return:

"Unspecified"

====================================================================
REAUTHORIZATION RULES
=====================

Extract continuation requirements exactly.

Preserve:

* AND logic
* OR logic

Ignore:

* initial authorization criteria
* refill frequency
* quantity limits
* dosing frequency

Examples:

positive clinical response

continued benefit

maintained disease control

improvement from baseline

If continuation approval exists but requirements are absent:

return:

"Unspecified"

====================================================================
IMPORTANT RULES
===============

1. Use ONLY provided text.

2. Do NOT use external knowledge.

3. Do NOT hallucinate.

4. Ignore unrelated drugs.

5. Ignore unrelated diseases.

6. Ignore unrelated indications.

7. Preserve policy meaning exactly.

8. Determine each output independently.


====================================================================
OUTPUT FORMAT
=============

Return ONLY valid JSON.

{{
"Initial Authorization Duration(in-months)": "<number | Unspecified>",
"Reauthorization Required": "<Yes | No>",
"Reauthorization Duration(in-months)": "<number | Unspecified | NA>",
"Reauthorization Requirements Documented in Policy": "<requirements | Unspecified | NA>"
}}

====================================================================
POLICY TEXT
===========

Initial Authorization retrieval chunk= {init_auth}
Reauthorization etrieval chunk= {reauth}

You may use information from ANY retrieval section.

A criterion may appear in a section retrieved for another attribute.

Do not restrict extraction to the section name.
Use all retrieved evidence when determining the final answer.
"""

CRITERIA_CONFIG = {

    "Eligibility": {

        "retrieval_prompts": {
            "age": AGE_RETRIEVAL_PROMPT,
            "tb": TB_TEST_RETRIEVAL_PROMPT,
            "specialist": SPECIALIST_TYPES_RETRIEVAL_PROMPT
        },

        "extraction_prompt": Eligibility_Extraction_prompt,

        "top_k": 5
    },

    "StepTherapy": {

        "retrieval_prompts": {
            "step": STEP_THERAPY_RETRIEVAL_PROMPT,
            "brand": BRANDED_STEP_COUNT_RETRIEVAL_PROMPT,
            "generic": GENERIC_STEP_COUNT_RETRIEVAL_PROMPT,
            "phototherapy": PHOTOTHERAPY_STEP_RETRIEVAL_PROMPT
        },

        "extraction_prompt": Step_Extraction_Prompt,

        "top_k": 5
    },

    "Authorization": {

        "retrieval_prompts": {
            "init_auth": INITIAL_AUTH_DURATION_RETRIEVAL_PROMPT,
            "reauth": REAUTHORIZATION_RETRIEVAL_PROMPT,
        },

        "extraction_prompt": Authorization_Extraction_Prompt,

        "top_k": 5
    },

    "Quantity": {

        "retrieval_prompts": {
            "quantity": QUANTITY_LIMITS_RETRIEVAL_PROMPT
        },

        "extraction_prompt": QUANTITY_LIMITS_EXTRACTION_PROMPT,

        "top_k": 5
    }
}


# =========================================================
# COMPLETE END-TO-END PAYER POLICY EXTRACTION PIPELINE
# =========================================================
#
# FLOW:
#
# PDF
#   ↓
# CHUNKING
#   ↓
# BM25 + DENSE INDEX
#   ↓
# HYBRID RETRIEVAL
#   ↓
# RERANKER
#   ↓
# LLM EXTRACTION
#   ↓
# FINAL STRUCTURED OUTPUT
#
# =========================================================


# =========================================================
# INSTALL
# =========================================================

# pip install pymupdf pdfplumber
# pip install sentence-transformers rank-bm25 faiss-cpu
# pip install groq


# =========================================================
# IMPORTS
# =========================================================





# =========================================================
# CHUNKER CONFIG
# =========================================================

MAX_CHARS = 2000
OVERLAP_CHARS=200


# =========================================================
# CLEANING
# =========================================================

def clean_text(text):

    text = re.sub(r'https?://\S+', ' ', text)

    text = re.sub(r'www\.\S+', ' ', text)

    text = re.sub(r'\s+', ' ', text)

    return text.strip()


def normalize_heading_for_removal(text):

    t = text.strip().lower()

    t = re.sub(r'\s+', '', t)

    t = re.sub(r'[^a-z0-9]', '', t)

    return t


def is_page_number(text):

    t = text.strip()

    patterns = [
        r'^page\s+\d+(\s+of\s+\d+)?$',
        r'^\d+\s+of\s+\d+$',
    ]

    for p in patterns:

        if re.match(p, t, re.I):
            return True

    return False


# =========================================================
# FONT HELPERS
# =========================================================

def is_bold(spans):

    bold_words = ["bold", "heavy", "black", "demi"]

    for s in spans:

        font = s.get("font", "").lower()

        if any(w in font for w in bold_words):
            return True

    return False


def is_underlined(spans, drawings):

    x0 = min(s["bbox"][0] for s in spans)
    x1 = max(s["bbox"][2] for s in spans)

    y0 = min(s["bbox"][1] for s in spans)
    y1 = max(s["bbox"][3] for s in spans)

    text_width = x1 - x0

    # ignore tiny text fragments
    if text_width < 30:
        return False

    for d in drawings:

        if d["type"] != "s":
            continue

        for item in d["items"]:

            if item[0] != "l":
                continue

            p1 = item[1]
            p2 = item[2]

            # horizontal only
            if abs(p1.y - p2.y) > 1:
                continue

            line_x0 = min(p1.x, p2.x)
            line_x1 = max(p1.x, p2.x)

            line_y = p1.y

            # must be directly below text
            if not (y1 <= line_y <= y1 + 2):
                continue

            overlap = (
                min(x1, line_x1)
                -
                max(x0, line_x0)
            )

            if overlap < text_width * 0.8:
                continue

            line_width = line_x1 - line_x0

            # reject long table borders
            if line_width > text_width * 1.5:
                continue

            # reject lines stretching across page
            if line_width > 300:
                continue

            # must start/end near text boundaries
            if abs(line_x0 - x0) > 10:
                continue

            if abs(line_x1 - x1) > 10:
                continue

            return True

    return False

def get_font_size(spans):

    return max(s["size"] for s in spans)


# =========================================================
# BODY FONT SIZE
# =========================================================

def get_body_font_size(pdf_path):

    doc = fitz.open(pdf_path)

    size_counter = Counter()

    for page in doc:

        data = page.get_text("dict")

        for block in data["blocks"]:

            if "lines" not in block:
                continue

            for line in block["lines"]:

                for span in line["spans"]:

                    sz = round(span["size"] * 2) / 2

                    size_counter[sz] += 1

    if not size_counter:
        return 10.0

    return size_counter.most_common(1)[0][0]


# =========================================================
# STRIKETHROUGH
# =========================================================

def is_strikethrough(line):

    if "spans" not in line:
        return False

    for span in line["spans"]:

        flags = span.get("flags", 0)

        if flags & 8:
            return True

    return False


# =========================================================
# HEADING DETECTION
# =========================================================

def is_heading(text, font_size, bold, underlined, body_size):

    t = text.strip()

    if len(t) <= 2:
        return False, 0

    if t in ["AND", "OR", "-OR-", "-AND-" ]:
        return False, 0

    score = 0

    if bold:
        score += 4

    if underlined:
      score+=3

    if font_size >= body_size + 2.0:
        score += 3

    elif font_size >= body_size + 1.0:
        score += 1

    if t.isupper() and len(t.split()) <= 8:
        score += 2

    if re.match(r'^[A-Z]\.', t):
        score += 1

    if re.match(r'^\d+\.', t):
        score += 1

    if len(t.split()) <= 8:
        score += 1

    if len(t.split()) > 18:
        score -= 3

    if t.endswith(".") or t.endswith(";"):
        score -= 3

    return score >= 3, score


# =========================================================
# REPEATED HEADINGS
# =========================================================

def collect_repeated_headings(pdf_path, body_size):

    doc = fitz.open(pdf_path)

    heading_counter = Counter()

    for page in doc:

        data = page.get_text("dict")

        for block in data["blocks"]:

            if "lines" not in block:
                continue

            for line in block["lines"]:

                spans = line["spans"]

                text = " ".join(
                    clean_text(s["text"])
                    for s in spans
                ).strip()

                if not text:
                    continue

                if is_page_number(text):
                    continue

                bold = is_bold(spans)

                font_size = get_font_size(spans)

                is_head,_=is_heading(
                    text,
                    font_size,
                    bold,
                    False,
                    body_size
                )
                if is_head:
                  heading_counter[text] += 1

    repeated = set()

    total_pages = len(doc)

    for h, c in heading_counter.items():

        if c >= max(3, total_pages * 0.4):

            repeated.add(h)

    return repeated


# =========================================================
# TABLE DETECTION
# =========================================================

def looks_like_real_table(
    table,
    page_height=None,
    bbox=None
):

    if not table or len(table) < 2:
        return False

    col_counts = []

    non_empty_rows = 0

    total_cell_len = 0

    cell_count = 0

    for row in table:

        cleaned = [
            clean_text(c or "")
            for c in row
        ]

        non_empty = sum(bool(x) for x in cleaned)

        if non_empty > 1:

            non_empty_rows += 1

            col_counts.append(non_empty)

        for c in cleaned:

            if c:

                total_cell_len += len(c)

                cell_count += 1

    if non_empty_rows < 2:
        return False

    avg_col_count = (
        sum(col_counts) / len(col_counts)
        if col_counts else 0
    )

    avg_cell_len = (
        total_cell_len / cell_count
        if cell_count else 0
    )

    max_cell_len = max(
        (
            len(clean_text(c or ""))
            for row in table
            for c in row
        ),
        default=0
    )

    unique_col_counts = len(set(col_counts))

    if page_height and bbox:

        height_coverage = (
            (bbox[3] - bbox[1]) / page_height
        )

        if (
            height_coverage > 0.60
            and
            avg_cell_len > 200
        ):
            return False

    if avg_col_count <= 2 and max_cell_len > 400:
        return False

    if unique_col_counts <= 2:
        return True

    return False


def get_table_boxes(page, page_height):

    boxes = []

    for t in page.find_tables():

        extracted = t.extract()

        if looks_like_real_table(
            extracted,
            page_height=page_height,
            bbox=t.bbox
        ):

            boxes.append(t.bbox)

    return boxes


def point_inside_table(x, y, boxes):

    for (x0, top, x1, bottom) in boxes:

        if (
            x0 - 3 <= x <= x1 + 3
            and
            top - 3 <= y <= bottom + 3
        ):
            return True

    return False


# =========================================================
# TABLE CHUNKS
# =========================================================

def extract_table_chunks(pdf_path):

    table_chunks = []

    with pdfplumber.open(pdf_path) as pdf:

        for page_num, page in enumerate(pdf.pages):

            page_height = page.height

            for t in page.find_tables():

                extracted = t.extract()

                if not looks_like_real_table(
                    extracted,
                    page_height=page_height,
                    bbox=t.bbox
                ):
                    continue

                table = extracted

                headers = None

                start_idx = 0

                for i, row in enumerate(table):

                    cleaned = [
                        clean_text(c or "")
                        for c in row
                    ]

                    if sum(bool(x) for x in cleaned) >= 2:

                        headers = cleaned

                        start_idx = i + 1

                        break

                if not headers:
                    continue

                for row in table[start_idx:]:

                    cleaned_row = [
                        clean_text(c or "")
                        for c in row
                    ]

                    if not any(cleaned_row):
                        continue

                    row_dict = {

                        h: v

                        for h, v in zip(headers, cleaned_row)

                        if h and v
                    }

                    if not row_dict:
                        continue

                    text = " | ".join(
                        f"{k}: {v}"
                        for k, v in row_dict.items()
                    )

                    table_chunks.append({

                        "type": "table_row",

                        "page": page_num + 1,

                        "headers": headers,

                        "row": row_dict,

                        "text": text
                    })

    return table_chunks


# =========================================================
# SAVE CHUNK
# =========================================================

def save_chunk(
    text_chunks,
    current_chunk,
    heading_stack,
    page_num
):

    if not current_chunk:
        return

    final_text = " ".join(current_chunk).strip()

    if not final_text:
        return

    text_chunks.append({

        "type": "text_chunk",

        "page": page_num + 1,

        "headings": [
            h["text"] if isinstance(h, dict) else h
            for h in heading_stack[-3:]
        ],

        "text": final_text
    })


# =========================================================
# TEXT EXTRACTION
# =========================================================

def extract_text_chunks(
    pdf_path,
    repeated_headings,
    body_size
):

    doc = fitz.open(pdf_path)

    text_chunks = []

    heading_stack = []

    current_chunk = []

    current_indent = None

    previous_heading_indent = None

    seen_lines = set()

    with pdfplumber.open(pdf_path) as plumber_pdf:

        for page_num in range(len(doc)):

            fitz_page = doc[page_num]

            plumber_page = plumber_pdf.pages[page_num]

            page_height = plumber_page.height

            table_boxes = get_table_boxes(
                plumber_page,
                page_height
            )

            data = fitz_page.get_text("dict")

            all_lines = []
            fitz_page = doc[page_num]
            drawings = fitz_page.get_drawings()

            for block in data["blocks"]:

                if "lines" not in block:
                    continue

                for line in block["lines"]:

                    if is_strikethrough(line):
                        continue

                    spans = line["spans"]

                    text = " ".join(
                        clean_text(s["text"])
                        for s in spans
                    ).strip()

                    if not text:
                        continue

                    if is_page_number(text):
                        continue

                    bbox = line["bbox"]

                    x0 = bbox[0]

                    y0 = (
                        bbox[1] + bbox[3]
                    ) / 2

                    if point_inside_table(
                        x0,
                        y0,
                        table_boxes
                    ):
                        continue

                    all_lines.append({

                        "text": text,

                        "x0": x0,

                        "y0": y0,

                        "spans": spans
                    })

            all_lines = sorted(
                all_lines,
                key=lambda z: (
                    round(z["y0"] / 3) * 3,
                    z["x0"]
                )
            )

            for item in all_lines:

                text = item["text"]

                if (
                    text in seen_lines
                    and
                    text in repeated_headings
                ):
                    continue

                seen_lines.add(text)

                spans = item["spans"]

                indent = item["x0"]

                bold = is_bold(spans)

                font_size = get_font_size(spans)
                underlined = is_underlined(spans, drawings)

                is_head, heading_score = is_heading(
                    text,
                    font_size,
                    bold,
                    underlined,
                    body_size
                )

                if is_head:

                    if text in repeated_headings:
                        continue

                    save_chunk(
                        text_chunks,
                        current_chunk,
                        heading_stack,
                        page_num
                    )

                    current_chunk = []

                    if len(heading_stack) == 0:

                        heading_stack.append({

                            "text": text,

                            "indent": indent,

                            "score": heading_score
                        })

                    else:

                        last = heading_stack[-1]

                        last_indent = last["indent"]

                        last_score = last["score"]

                        # ==================================
                        # CHILD HEADING
                        # ==================================

                        if indent > last_indent + 10:

                            heading_stack.append({

                                "text": text,

                                "indent": indent,

                                "score": heading_score
                            })

                        # ==================================
                        # SAME INDENT
                        # ==================================

                        elif abs(indent - last_indent) < 10:

                            # sibling

                            if abs(
                                heading_score
                                -
                                last_score
                            ) <= 1:

                                heading_stack.pop()

                                heading_stack.append({

                                    "text": text,

                                    "indent": indent,

                                    "score": heading_score
                                })

                            # weaker heading

                            elif heading_score < last_score:

                                heading_stack.append({

                                    "text": text,

                                    "indent": indent,

                                    "score": heading_score
                                })

                            # stronger heading

                            else:

                                while (

                                    heading_stack

                                    and

                                    heading_stack[-1]["score"]
                                    <
                                    heading_score
                                ):

                                    heading_stack.pop()

                                heading_stack.append({

                                    "text": text,

                                    "indent": indent,

                                    "score": heading_score
                                })

                        # ==================================
                        # LESS INDENT
                        # ==================================

                        else:

                            while (

                                heading_stack

                                and

                                heading_stack[-1]["indent"]
                                >=
                                indent
                            ):

                                heading_stack.pop()

                            heading_stack.append({

                                "text": text,

                                "indent": indent,

                                "score": heading_score
                            })

                    heading_stack = heading_stack[-5:]

                    previous_heading_indent = indent

                    current_indent = None

                    continue

                if current_indent is None:

                    current_indent = indent
                    logical_connectors = {
                                          "OR",
                                          "-OR-",
                                          "AND",
                                          "-AND-"
                                         }

                if text.strip().upper() in logical_connectors:
                    current_chunk.append(text)
                    continue

                if (
                    abs(indent - current_indent) < 25
                    or
                    indent > current_indent
                ):

                    current_chunk.append(text)

                else:

                    save_chunk(
                        text_chunks,
                        current_chunk,
                        heading_stack,
                        page_num
                    )

                    current_chunk = [text]

                    current_indent = indent

                if sum(
                    len(x)
                    for x in current_chunk
                ) > MAX_CHARS:

                    full_chunk_text="".join(current_chunk)
                    overlap_text=full_chunk_text[-OVERLAP_CHARS:]
                    save_chunk(
                        text_chunks,
                        current_chunk,
                        heading_stack,
                        page_num
                    )

                    current_chunk = [overlap_text]

                    current_indent = indent

    save_chunk(
        text_chunks,
        current_chunk,
        heading_stack,
        len(doc) - 1
    )

    return text_chunks


# =========================================================
# REMOVE REFERENCES
# =========================================================

def remove_reference_chunks(all_chunks):

    filtered = []

    bad_headings = {

        "reference",
        "references",
        "source",
        "sources",
        "bibliography",
        "citations",
        "appendix",
        "appendices",
        "acknowledgements",
        "acknowledgments",
        "documenthistory",
        "changehistory",
        "additionalresources"
    }

    for chunk in all_chunks:

        if chunk["type"] != "text_chunk":

            filtered.append(chunk)

            continue

        headings = chunk.get("headings", [])

        if not headings:

            filtered.append(chunk)

            continue

        latest_heading = normalize_heading_for_removal(
            headings[-1]
        ).lower().strip()

        if latest_heading in bad_headings:
            continue

        filtered.append(chunk)

    return filtered


# =========================================================
# PROCESS PDF
# =========================================================

def process_pdf(pdf_path):

    print(f"Processing: {pdf_path}")

    body_size = get_body_font_size(pdf_path)

    repeated = collect_repeated_headings(
        pdf_path,
        body_size
    )

    table_chunks = extract_table_chunks(
        pdf_path
    )

    text_chunks = extract_text_chunks(
        pdf_path,
        repeated,
        body_size
    )

    all_chunks = table_chunks + text_chunks

    all_chunks = remove_reference_chunks(
        all_chunks
    )

    print("FINAL CHUNKS:", len(all_chunks))

    return all_chunks


# =========================================================
# BUILD RETRIEVAL INDEX
# =========================================================

def build_retrieval_pipeline(chunks):

    print("Loading embedding model...")
    device = "cuda" if torch.cuda.is_available() else "cpu"

    embedding_model = SentenceTransformer(
        "BAAI/bge-large-en-v1.5",
        device=device
    )

    print("Loading reranker...")

    reranker = CrossEncoder(
        "BAAI/bge-reranker-base",
        device=device
    )

    documents = []

    for chunk in chunks:

        if chunk["type"] == "table_row":

            text = chunk["text"]

        else:

            heading_text = " > ".join(
                chunk["headings"]
            )

            text = (
                f"HEADINGS: {heading_text}\n\n"
                f"CONTENT:\n{chunk['text']}"
            )

        documents.append(text)

    def tokenize(text):

        text = text.lower()

        text = re.sub(
            r'[^a-z0-9\s]',
            ' ',
            text
        )

        return text.split()

    tokenized_docs = [
        tokenize(doc)
        for doc in documents
    ]

    bm25 = BM25Okapi(tokenized_docs)

    print("Creating embeddings...")

    dense_embeddings = embedding_model.encode(
        documents,
        normalize_embeddings=True,
        batch_size=16,
        show_progress_bar=True
    )

    dense_embeddings = np.array(
        dense_embeddings,
        dtype="float32"
    )

    dimension = dense_embeddings.shape[1]

    index = faiss.IndexFlatIP(dimension)

    index.add(dense_embeddings)

    return {

        "chunks": chunks,

        "documents": documents,

        "embedding_model": embedding_model,

        "reranker": reranker,

        "bm25": bm25,

        "index": index,

        "tokenize": tokenize
    }


# =========================================================
# HYBRID SEARCH
# =========================================================

def hybrid_search(

    retrieval_system,

    query,

    top_k_dense=20,

    top_k_sparse=20,

    final_top_k=10,

    sparse_weight=0.65,

    dense_weight=0.35
):

    bm25 = retrieval_system["bm25"]

    index = retrieval_system["index"]

    embedding_model = retrieval_system["embedding_model"]

    reranker = retrieval_system["reranker"]

    documents = retrieval_system["documents"]

    chunks = retrieval_system["chunks"]

    tokenize = retrieval_system["tokenize"]

    tokenized_query = tokenize(query)

    bm25_scores = bm25.get_scores(
        tokenized_query
    )

    bm25_top_idx = np.argsort(
        bm25_scores
    )[::-1][:top_k_sparse]

    query_embedding = embedding_model.encode(
        [query],
        normalize_embeddings=True
    )

    query_embedding = np.array(
        query_embedding,
        dtype="float32"
    )

    dense_scores, dense_idx = index.search(
        query_embedding,
        top_k_dense
    )

    dense_scores = dense_scores[0]

    dense_idx = dense_idx[0]

    combined_scores = {}

    bm25_max = max(bm25_scores) + 1e-6

    for idx in bm25_top_idx:

        normalized = bm25_scores[idx] / bm25_max

        combined_scores[idx] = (
            sparse_weight * normalized
        )

    for idx, score in zip(
        dense_idx,
        dense_scores
    ):

        if idx not in combined_scores:
            combined_scores[idx] = 0

        combined_scores[idx] += (
            dense_weight * float(score)
        )

    merged = sorted(
        combined_scores.items(),
        key=lambda x: x[1],
        reverse=True
    )

    candidate_indices = [
        x[0]
        for x in merged[:40]
    ]

    rerank_pairs = [
        (query, documents[idx])
        for idx in candidate_indices
    ]

    rerank_scores = reranker.predict(
        rerank_pairs
    )

    reranked = list(zip(
        candidate_indices,
        rerank_scores
    ))

    reranked = sorted(
        reranked,
        key=lambda x: x[1],
        reverse=True
    )

    final_results = []

    for idx, score in reranked[:final_top_k]:

        final_results.append({

            "score": float(score),

            "chunk": chunks[idx]
        })

    return final_results


# =========================================================
# GROQ CLIENT
# =========================================================

client = Groq(
    api_key=GROQ_API_KEY
)
# client = genai.Client(
#     api_key=GEMINI_API_KEY
#   )


# =========================================================
# LLM CALL
# =========================================================

def run_llm_extraction(prompt):

    response = client.chat.completions.create(

        model="llama-3.3-70b-versatile",

        temperature=0,

        messages=[

            {
                "role": "system",

                "content": (
                    "You are a highly accurate medical payer policy extraction system. "
                    "You extract information conservatively and only from provided text."
                )
            },

            {
                "role": "user",


                "content": prompt
            }
        ]
    )

    return response.choices[0].message.content
      # response = client.models.generate_content(
      #     model="gemini-2.5-flash",
      #     contents=prompt,
      #     config={
      #       "temperature": 0
      #             }
      #   )

      # return response.text


# =========================================================
# RETRIEVE CONTEXT
# =========================================================

def retrieve_context(

    retrieval_system,

    retrieval_prompt,

    top_k
):

    results = hybrid_search(

        retrieval_system,

        retrieval_prompt,

        final_top_k=top_k
    )

    retrieved_chunks = []

    for r in results:

        chunk = r["chunk"]

        if chunk["type"] == "table_row":

            text = chunk["text"]

        else:

            headings = " > ".join(
                chunk["headings"]
            )

            text = (
                f"HEADINGS: {headings}\n\n"
                f"{chunk['text']}"
            )

        retrieved_chunks.append(text)

    context = "\n\n".join(
        retrieved_chunks
    )

    return context


# =========================================================
# PARAMETER EXTRACTION
# =========================================================

def extract_parameter(

    retrieval_system,

    parameter_name,

    config,

    drug_name
):

    print(f"\nEXTRACTING: {parameter_name}")

    # =====================================
    # RETRIEVE ALL CONTEXTS
    # =====================================

    contexts = {}

    global_seen = set()

    for context_name, retrieval_prompt in config[
        "retrieval_prompts"
    ].items():

        prompt = retrieval_prompt.format(
            DRUG_NAME=drug_name
        )

        context = retrieve_context(

            retrieval_system,

            prompt,

            config["top_k"]
        )

        # =====================================
        # DEDUPLICATE CHUNKS
        # =====================================

        unique_chunks = []

        for chunk in context.split("\n\n"):

            normalized = re.sub(
                r"\s+",
                " ",
                chunk.strip()
            )

            if not normalized:
                continue

            if normalized in global_seen:
                continue

            global_seen.add(normalized)

            unique_chunks.append(chunk)

        contexts[context_name] = "\n\n".join(
            unique_chunks
        )

    # =====================================
    # BUILD EXTRACTION PROMPT
    # =====================================

    format_dict = {

        "DRUG_NAME": drug_name
    }

    for k, v in contexts.items():

        format_dict[k] = v

    llm_prompt = config[
        "extraction_prompt"
    ].format(**format_dict)

    # Optional debugging
    print(
        f"Prompt length: {len(llm_prompt):,} chars"
    )

    # =====================================
    # LLM
    # =====================================

    output = run_llm_extraction(
        llm_prompt
    )
    time.sleep(20)

    return output

# =========================================================
# MAIN PIPELINE
# =========================================================

def run_full_pipeline(

    pdf_path,

    drug_name,

    criteria_config
):

    # =====================================
    # CHUNKING
    # =====================================

    chunks = process_pdf(
        pdf_path
    )

    # =====================================
    # RETRIEVAL SYSTEM
    # =====================================

    retrieval_system = build_retrieval_pipeline(
        chunks
    )

    # =====================================
    # FINAL RESULTS
    # =====================================

    final_results = {}

    # =====================================
    # LOOP THROUGH PARAMETERS
    # =====================================

    for parameter_name, config in criteria_config.items():

        try:

            result = extract_parameter(

                retrieval_system,

                parameter_name,

                config,

                drug_name
            )

            final_results[
                parameter_name
            ] = result

        except Exception as e:

            final_results[
                parameter_name
            ] = f"ERROR: {str(e)}"

    return final_results







# =========================================================
# RUN PIPELINE
# =========================================================




# =========================================================
# PRINT RESULTS
# =========================================================

# print("\n" + "=" * 120)
# print("FINAL EXTRACTION RESULTS")
# print("=" * 120)

# for k, v in results.items():

#     print("\n" + "-" * 100)

#     print(k.upper())

#     print("-" * 100)

#     print(v)



# ==========================================================
# LOAD WORKBOOK
# ==========================================================

wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

# ==========================================================
# COLUMN MAPPING
# ==========================================================

COLUMN_MAP = {
    "Age": "C",
    "Step Therapy Requirements Documented in Policy": "D",
    "Number of Steps through Brands": "E",
    "Number of Steps through Generic": "F",
    "Step through-Phototherapy": "G",
    "TB Test required": "H",
    "Quantity Limits": "I",
    "Specialist Types": "J",
    "Initial Authorization Duration(in-months)": "K",
    "Reauthorization Duration(in-months)": "L",
    "Reauthorization Required": "M",
    "Reauthorization Requirements Documented in Policy": "N"
}

# ==========================================================
# LOOP THROUGH ALL SUBMISSIONS
# ==========================================================

for row in range(2,ws.max_row + 1):

    if ws[f"A{row}"].value is None:
      break

    pdf_name = ws[f"A{row}"].value
    drug_name = ws[f"B{row}"].value

    if not pdf_name or not drug_name:
        continue

    print(f"\nProcessing Row {row}")
    print(f"PDF: {pdf_name}")
    print(f"Drug: {drug_name}")

    pdf_path = f"{PDF_FOLDER}/{pdf_name}"

    try:

        results = run_full_pipeline(
            pdf_path=pdf_path,
            drug_name=drug_name,
            criteria_config=CRITERIA_CONFIG
        )

        # ==================================================
        # MERGE ALL JSON OUTPUTS
        # ==================================================

        merged_json = {}

        for criterion_name, output in results.items():

            if isinstance(output, str):

                try:
                    # extract JSON even if LLM wrapped it in text/markdown
                    match = re.search(r'\{.*\}', output, re.DOTALL)

                    if not match:
                        continue

                    parsed = json.loads(match.group())

                except Exception as e:

                    print(f"Failed parsing {criterion_name}")
                    print(output[:500])

                    continue

            elif isinstance(output, dict):

                parsed = output

            else:

                continue

            merged_json.update(parsed)

        # ==================================================
        # WRITE TO EXCEL
        # ==================================================

        field_mapping = {
            "Age": merged_json.get("Age"),
            "Step Therapy Requirements Documented in Policy":
                merged_json.get("Step Therapy Requirements Documented in Policy"),

            "Number of Steps through Brands":
                merged_json.get("Number of Steps through Brands"),

            "Number of Steps through Generic":
                merged_json.get("Number of Steps through Generic"),

            "Step through-Phototherapy":
                merged_json.get("Step through-Phototherapy"),

            "TB Test required":
                merged_json.get("TB Test required"),

            "Quantity Limits":
                merged_json.get("Quantity Limits"),

            "Specialist Types":
                merged_json.get("Specialist Types"),

            "Initial Authorization Duration(in-months)":
                merged_json.get("Initial Authorization Duration(in-months)"),

            "Reauthorization Duration(in-months)":
                merged_json.get("Reauthorization Duration(in-months)"),

            "Reauthorization Required":
                merged_json.get("Reauthorization Required"),

            "Reauthorization Requirements Documented in Policy":
                merged_json.get("Reauthorization Requirements Documented in Policy")
        }

        for field_name, value in field_mapping.items():

            col = COLUMN_MAP[field_name]

            ws[f"{col}{row}"] = (
                json.dumps(value)
                if isinstance(value, (dict, list))
                else value
            )

        wb.save(EXCEL_PATH)

        print(f"Completed Row {row}")

    except Exception as e:

        print(f"Error in row {row}")
        print(str(e))

        ws[f"O{row}"] = str(e)

        wb.save(EXCEL_PATH)

# ==========================================================
# FINAL SAVE
# ==========================================================

wb.save(EXCEL_PATH)

print("\nAll rows processed.")



# Access Score

# ==========================================================
# ACCESS SCORE CALCULATION
# ==========================================================


# These dict. were calculated seperately from training data

AGE_DIST= {18: 27, 6: 22, 12: 1, 4: 1}
BRAND_STEPS_DIST= {0: 21, 1: 42, 2: 6, 3: 2}
GENERIC_STEPS_DIST= {0: 23, 1: 46, 2: 8, 3: 2}
INITIAL_AUTH_DIST= {12: 15, 6: 14, 1: 1, 3: 4, 4: 3}
REAUTH_DIST= {12: 34, 0: 5, 6: 2}

BINARY_SCORES = {
    "TB Test required": {
        "Yes": 60.49,
        "No": 39.51
    },
    "Quantity Limits": {
        "Yes": 79.03,
        "No": 20.97
    },
    "Reauthorization Required": {
        "Yes": 7.89,
        "No": 92.11
    },
    "Step through-Phototherapy": {
        "Yes": 85.19,
        "No": 14.81
    },
    "Specialist Types": {
       "Yes": 45,
       "No": 34
    }
}

WEIGHTS = {
    "BrandSteps": 0.25,
    "GenericSteps": 0.15,
    "Phototherapy": 0.10,
    "QuantityLimits": 0.10,
    "ReauthRequired": 0.05,
    "InitialAuth": 0.10,
    "ReauthDuration": 0.05,
    "TB": 0.05,
    "Specialist": 0.10,
    "Age": 0.05
}

BIOLOGICS = {
    "TREMFYA","STELARA","SKYRIZI","COSENTYX",
    "BIMZELX","ILUMYA","SILIQ",
    "ENBREL","REMICADE","CIMZIA"
}

BIOSIMILARS = {
    "AMJEVITA","YESINTEK","OTULFI"
}

NON_BIOLOGICS = {
    "OTEZLA","ACITRETIN"
}


def extract_numeric(value):

    if value is None:
        return np.nan

    text = str(value).strip().lower()

    if text in ["", "na"]:
        return 0
    if text in ["not specified","unspecified"]:
        return np.nan

    nums = re.findall(r'\d+', text)

    if nums:
        return int(nums[0])

    return np.nan

def extract_age(age_text):

    if age_text is None:
        return np.nan

    text = str(age_text).strip().lower()

    if text in ["not specified", "na", "n/a", "none", ""]:
        return np.nan

    if "fda labelled age" in text:
        return 6       #gives a good score as you can't be better than this       

    m = re.search(r'>=\s*(\d+)', text)
    if m:
        return int(m.group(1))

    nums = re.findall(r'\d+', text)

    if nums:
        return int(nums[0])

    return np.nan

def specialist_to_binary(value):

    if value is None:
        return None

    text = str(value).strip().lower()

    if text in [
        "",
        "not specified",
        "unspecified",
        "unknown",
        "none"
    ]:
        return None

    if text in [
        "no specialist required",
        "no specialist"
        "na"
    ]:
        return "No"

    return "Yes"

def clean_binary(value):

    if value is None:
        return None

    text = str(value).strip().lower()

    # YES variants
    if text in [
        "yes",
        "y",
        "true"
    ]:
        return "Yes"

    # NO variants
    if text in [
        "no",
        "n",
        "false",
        "na",
        "n/a",
        "not applicable",
        "not required"
    ]:
        return "No"

    # Unknown values
    if text in [
        "",
        "none",
        "null",
        "not specified",
        "unspecified",
        "unknown"
    ]:
        return None

    return "Yes"


def lower_better_score(value, distribution):

    total = sum(distribution.values())

    better = sum(
        count
        for k, count in distribution.items()
        if k >= value
    )

    return (better / total) * 100


def higher_better_score(value, distribution):

    total = sum(distribution.values())

    better = sum(
        count
        for k, count in distribution.items()
        if k <= value
    )

    return (better / total) * 100


def bucket_score(score):

    if score < 12.5:
        return 0

    elif score < 37.5:
        return 25

    elif score < 62.5:
        return 50

    elif score < 87.5:
        return 75

    return 100


# ==========================================================
# SCORE EVERY ROW
# ==========================================================

for row in range(2, ws.max_row + 1):
    
    if ws[f"A{row}"].value is None:
      break

    weighted_score = 0

    # --------------------------------------------------
    # AGE (COLUMN C)
    # --------------------------------------------------

    age = extract_age(ws[f"C{row}"].value)

    if pd.isna(age):
        age_score = 50
    else:
        age_score = lower_better_score(
            age,
            AGE_DIST
        )

    weighted_score += (
        age_score * WEIGHTS["Age"]
    )

    # --------------------------------------------------
    # BRAND STEPS (COLUMN E)
    # --------------------------------------------------

    brand_steps = extract_numeric(
        ws[f"E{row}"].value
    )

    if pd.isna(brand_steps):
        brand_score = 50
    else:
        brand_score = lower_better_score(
            brand_steps,
            BRAND_STEPS_DIST
        )

    weighted_score += (
        brand_score * WEIGHTS["BrandSteps"]
    )

    # --------------------------------------------------
    # GENERIC STEPS (COLUMN F)
    # --------------------------------------------------

    generic_steps = extract_numeric(
        ws[f"F{row}"].value
    )

    if pd.isna(generic_steps):
        generic_score = 50
    else:
        generic_score = lower_better_score(
            generic_steps,
            GENERIC_STEPS_DIST
        )

    weighted_score += (
        generic_score * WEIGHTS["GenericSteps"]
    )

    # --------------------------------------------------
    # INITIAL AUTH (COLUMN K)
    # --------------------------------------------------

    initial_auth = extract_numeric(
        ws[f"K{row}"].value
    )

    if pd.isna(initial_auth):
        initial_auth_score = 50
    else:
        initial_auth_score = higher_better_score(
            initial_auth,
            INITIAL_AUTH_DIST
        )

    weighted_score += (
        initial_auth_score * WEIGHTS["InitialAuth"]
    )

    # --------------------------------------------------
    # REAUTH DURATION (COLUMN L)
    # --------------------------------------------------

    reauth_duration = extract_numeric(
        ws[f"L{row}"].value
    )

    if pd.isna(reauth_duration):
        reauth_duration_score = 50
    else:
        reauth_duration_score = higher_better_score(
            reauth_duration,
            REAUTH_DIST
        )

    weighted_score += (
        reauth_duration_score
        * WEIGHTS["ReauthDuration"]
    )

    # --------------------------------------------------
    # PHOTOTHERAPY (COLUMN G)
    # --------------------------------------------------

    photo = clean_binary(
        ws[f"G{row}"].value
    )

    if photo in ["Yes", "No"]:
        photo_score = BINARY_SCORES[
            "Step through-Phototherapy"
        ][photo]
    else:
        photo_score = 50

    weighted_score += (
        photo_score * WEIGHTS["Phototherapy"]
    )

    # --------------------------------------------------
    # TB TEST (COLUMN H)
    # --------------------------------------------------

    tb = clean_binary(
        ws[f"H{row}"].value
    )

    if tb in ["Yes", "No"]:
        tb_score = BINARY_SCORES[
            "TB Test required"
        ][tb]
    else:
        tb_score = 50

    weighted_score += (
        tb_score * WEIGHTS["TB"]
    )

    # --------------------------------------------------
    # QUANTITY LIMITS (COLUMN I)
    # --------------------------------------------------

    qty = clean_binary(
        ws[f"I{row}"].value
    )

    if qty in ["Yes", "No"]:
        qty_score = BINARY_SCORES[
            "Quantity Limits"
        ][qty]
    else:
        qty_score = 50

    weighted_score += (
        qty_score * WEIGHTS["QuantityLimits"]
    )

    # --------------------------------------------------
    # SPECIALIST (COLUMN J)
    # --------------------------------------------------

    specialist = specialist_to_binary(
        ws[f"J{row}"].value
    )

    if specialist in ["Yes", "No"]:
        specialist_score = BINARY_SCORES[
            "Specialist Types"
        ][specialist]
    else:
        specialist_score = 50

    weighted_score += (
        specialist_score
        * WEIGHTS["Specialist"]
    )

    # --------------------------------------------------
    # REAUTH REQUIRED (COLUMN M)
    # --------------------------------------------------

    reauth_required = clean_binary(
        ws[f"M{row}"].value
    )

    if reauth_required in ["Yes", "No"]:
        reauth_required_score = BINARY_SCORES[
            "Reauthorization Required"
        ][reauth_required]
    else:
        reauth_required_score = 50

    weighted_score += (
        reauth_required_score
        * WEIGHTS["ReauthRequired"]
    )

    # --------------------------------------------------
    # DRUG PENALTY
    # --------------------------------------------------

    drug_name = str(
        ws[f"B{row}"].value
    ).strip().upper()

    if drug_name in BIOSIMILARS:
        weighted_score *= 0.90

    elif drug_name in NON_BIOLOGICS:
        weighted_score *= 0.80

    # --------------------------------------------------
    # FINAL BUCKET
    # --------------------------------------------------

    final_score = bucket_score(
        weighted_score
    )

    ws[f"O{row}"] = final_score

    print(
        f"Row {row} | "
        f"{drug_name} | "
        f"Raw={round(weighted_score,2)} | "
        f"Final={final_score}"
    )

# ==========================================================
# SAVE
# ==========================================================

wb.save(EXCEL_PATH)

print("Access Scores Completed")
